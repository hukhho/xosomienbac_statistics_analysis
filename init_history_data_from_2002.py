import openpyxl
from datetime import datetime
import pyodbc

# Replace 'your_file.xlsx' with the name of your Excel file
file_path = 'data.xlsx'

# Load the workbook and select the sheet
workbook = openpyxl.load_workbook(file_path)

# Read data from the excel sheet
sheet_names = workbook.sheetnames
for sheet_name in sheet_names:
    print(f"sheet name : " + sheet_name)
    
def is_valid_cell_value(value):
    if isinstance(value, str) and len(value) == 5 and value.isdigit():
        return True
    return False

    
def save_to_sql_server(result):
    # Connect to SQL Server
    connection_string = "DRIVER={ODBC Driver 17 for SQL Server};SERVER=DESKTOP-SJS8V3V;DATABASE=hunglotto;UID=sa;PWD=123"
    conn = pyodbc.connect(connection_string)
    cursor = conn.cursor()

    # Insert data into SQL Server
    for date, value in result:
        # Parse date in the format dd-MM-yyyy
        parsed_date = datetime.strptime(date, '%d-%m-%Y')
        
        #Insert the data into the database
        query = f"INSERT INTO lotto (date_column, value_column) VALUES ('{parsed_date}', '{value}')"
        cursor.execute(query)
        
    # cursor.execute("select * from lotto")
    
    # results = cursor.fetchall()
    
    # Commit the changes and close the connection
    conn.commit()
    cursor.close()
    conn.close()
    
def read_sheet(workbook):
    sheet_names = workbook.sheetnames
    result = []
    for sheet_name in sheet_names:
        data = []
        sheet = workbook[sheet_name]
        
        for row in range(2, sheet.max_row + 1):
            row_data = []
            for col in range(2, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None and is_valid_cell_value(cell_value):
                    row_data.append(cell_value)
                else:
                    row_data.append("NULL")
            data.append(row_data)
            # Continue with the previous code to convert the data into the desired format
        
        year = sheet_name
        print(f"year : " + year)

        for row_idx, row in enumerate(data):
            for col_idx, value in enumerate(row):
                date_str = f"{row_idx + 1:02d}-{col_idx + 1:02d}-{year}"
                
                # Validate the date format before appending
                try:
                    datetime.strptime(date_str, "%d-%m-%Y")
                    result.append((date_str, value))
                except ValueError as e:
                    print(f"Invalid date: {date_str} + value: {value} ({e})")
                    continue

        # Sort the result by the real date
        result.sort(key=lambda x: datetime.strptime(x[0], "%d-%m-%Y"))
    return result;
            
result = read_sheet(workbook=workbook)

# for date, value in result:
#     print(date, value)
# for res in result:
#     if res. == "NULL":
#         result.remove(res)
save_to_sql_server(result)

from statistics import mean

def filter_date(target_date):
    target_date_obj = datetime.strptime(target_date, "%d-%m-%Y")

    # Filter the result list to only include dates before the target_date
    data_before_target_date = [(date, value) for date, value in result if datetime.strptime(date, "%d-%m-%Y") < target_date_obj]

    return data_before_target_date

def predict_with_date(target_date):
    # Calculate the average value for '14-06-2023'
    data_before_target_date = filter_date(target_date)


    target_day = int(target_date.split('-')[0])

    values_on_target_day = []

    for date, value in data_before_target_date:
        day = int(date.split('-')[0])
        if day == target_day and value != "NULL":
            values_on_target_day.append(float(value))  # Convert value to float

    if values_on_target_day:
        predicted_value = mean(values_on_target_day)
        # print(f"Predicted value for {target_date}: {predicted_value}")
        return predicted_value;
    else:
        return None;
        # print(f"No data available to predict the value for {target_date}")


# ... (previous code)
def get_last_two_digits(value):
    value_as_int = int(value)
    value_str = str(value_as_int)
    last_two_digits = value_str[-2:]
    return int(last_two_digits)

                          
                                     
def calculate_profit_loss(historical_data, bet_value, bet_amount, return_rate):
    total_profit = 0
    
    for date, value in historical_data:
        if value != "NULL" and get_last_two_digits(float(value)) == bet_value:
            win_amount = bet_amount * return_rate
            #print(f"Winning with betting {bet_value} on {value} in {date}: +{win_amount}")
            total_profit += win_amount
        elif value != "NULL":
            #print(f"Loss with betting {bet_value} on {value} in {date}: -{bet_amount}")
            total_profit -= bet_amount

    return total_profit

# Example usage of the calculate_profit_loss function
# bet_amount = 1
# return_rate = 97

# profit_loss = calculate_profit_loss(result, bet_value, bet_amount, return_rate)
# print(f"Total profit/loss for betting on '{bet_value}' with bet amount '${bet_amount}' and return rate '{return_rate}' x: ${profit_loss}")    

# ans = predict_with_date('09-06-2023')
# print(get_last_two_digits(ans))


def calculate_profit_loss_with_prediction(historical_data, bet_amount, return_rate):
    total_profit = 0
    
    # Get the predicted value using predict_with_date function
   
    for date, value in historical_data:

        predict = predict_with_date(date)
        if predict != None:
            predicted_value = get_last_two_digits(predict)
            if predicted_value != None:
                if value != "NULL":
                    result = get_last_two_digits(float(value))
                    if value != "NULL" and predicted_value == result:
                        win_amount = bet_amount * return_rate
                        print(f"Winning with betting {predicted_value} on {result} in {date}: +{win_amount}")
                        total_profit += win_amount
                    elif value != "NULL":
                        print(f"Loss with betting {predicted_value} on {result} in {date}: -{bet_amount}")
                        total_profit -= bet_amount
            else:
                print(f"Not betting on {date} because not enough data")
    return total_profit

# profit_loss_with_prediction = calculate_profit_loss_with_prediction(result, bet_amount, return_rate)
# print(f"Total profit/loss with prediction for betting  with bet amount '${bet_amount}' and return rate '{return_rate}' x: ${profit_loss_with_prediction}")